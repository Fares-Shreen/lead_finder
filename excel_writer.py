import os
from datetime import datetime

from openpyxl import Workbook, load_workbook

from config import OUTPUT_DIR

HEADERS = ["Company Name", "Field", "Location", "Website", "LinkedIn",
           "Email(s)", "Phone(s)", "Source", "Source URL", "Found At"]


def _output_path(filename=None):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    # Single persistent file — every search run (today, tomorrow, next month)
    # appends to this same sheet until you swap dedupe.py for a real database.
    filename = filename or "leads_master.xlsx"
    return os.path.join(OUTPUT_DIR, filename)


def append_companies(records, filename=None):
    """
    records: list of dicts with keys name, field, location, website,
             linkedin, emails (list), phones (list), source
    Appends to an existing workbook if present, otherwise creates one.
    Returns the file path written to.
    """
    path = _output_path(filename)

    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"
        ws.append(HEADERS)

    for rec in records:
        ws.append([
            rec.get("name", ""),
            rec.get("field", ""),
            rec.get("location", ""),
            rec.get("website", "") or "",
            rec.get("linkedin", "") or "",
            ", ".join(rec.get("emails", []) or []),
            ", ".join(rec.get("phones", []) or []),
            rec.get("source", ""),
            rec.get("source_url", ""),
            datetime.now().strftime("%Y-%m-%d %H:%M"),
        ])

    for i, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[chr(64 + i)].width = max(16, len(header) + 4)

    wb.save(path)
    return path